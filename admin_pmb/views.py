import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone

from datetime import date, timedelta
from django.db.models import Sum, Count, Q
from pembayaran.models import Tagihan, KonfirmasiPembayaran

from pendaftaran.models import Pendaftaran, ProfilPendaftar
from dokumen.models import DokumenPendaftar
from master.models import JalurPenerimaan, GelombangPenerimaan, PengaturanSistem
from chatbot.models import KnowledgeBase
from accounts.views import WARNA_FAKULTAS, warna_fakultas
logger = logging.getLogger(__name__)


def cek_admin(user):
    return user.is_authenticated and user.role in [
        'admin_pmb', 'operator_pmb', 'panitia_seleksi', 'pimpinan'
    ]


def get_sidebar_counts():
    """Hitung badge notifikasi untuk sidebar"""
    return {
        'pending_verifikasi': Pendaftaran.objects.filter(status='LENGKAP').count(),
        'pending_dokumen': DokumenPendaftar.objects.filter(status_verifikasi='menunggu').count(),
    }


@login_required
def dashboard(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    # Statistik utama
    total        = Pendaftaran.objects.count()
    hari_ini     = Pendaftaran.objects.filter(
        tgl_daftar__date=timezone.now().date()).count()
    perlu_verif  = Pendaftaran.objects.filter(status='LENGKAP').count()
    dok_pending  = DokumenPendaftar.objects.filter(status_verifikasi='menunggu').count()
    lulus        = Pendaftaran.objects.filter(status='LULUS_SELEKSI').count()
    daftar_ulang = Pendaftaran.objects.filter(status='DAFTAR_ULANG').count()
    draft        = Pendaftaran.objects.filter(status='DRAFT').count()
    tidak_lulus  = Pendaftaran.objects.filter(status='TIDAK_LULUS').count()

    # Per jalur
    per_jalur = Pendaftaran.objects.values(
        'jalur__nama_jalur'
    ).annotate(total=Count('id')).order_by('-total')

    # Per prodi (+ warna per fakultas)
    from accounts.views import warna_fakultas
    per_prodi_qs = Pendaftaran.objects.values(
        'prodi_pilihan_1__nama_prodi',
        'prodi_pilihan_1__kode_fakultas',
    ).annotate(total=Count('id')).order_by('-total')[:8]

    # Siapkan data untuk chart + legend
    per_prodi = []
    warna_prodi_list = []
    for p in per_prodi_qs:
        kode_fak = p['prodi_pilihan_1__kode_fakultas']
        w = warna_fakultas(kode_fak)
        per_prodi.append({
            'prodi_pilihan_1__nama_prodi': p['prodi_pilihan_1__nama_prodi'],
            'total': p['total'],
            'kode_fakultas': kode_fak,
            'warna': w['text'],
        })
        warna_prodi_list.append(w['text'])

    # Legend fakultas — hanya yang muncul di chart
    from accounts.views import WARNA_FAKULTAS
    kode_fakultas_muncul = {p['kode_fakultas'] for p in per_prodi if p['kode_fakultas']}
    legend_fakultas = [
        {'kode': k, 'hex': WARNA_FAKULTAS[k]['text']}
        for k in kode_fakultas_muncul if k in WARNA_FAKULTAS
    ]

    # 10 terbaru
    terbaru = Pendaftaran.objects.select_related(
        'user', 'jalur', 'prodi_pilihan_1', 'gelombang'
    ).order_by('-tgl_daftar')[:10]

    pengaturan = PengaturanSistem.get()

    context = {
        'total': total, 'hari_ini': hari_ini,
        'perlu_verif': perlu_verif, 'dok_pending': dok_pending,
        'lulus': lulus, 'daftar_ulang': daftar_ulang,
        'draft': draft, 'tidak_lulus': tidak_lulus,
        'per_jalur': per_jalur, 'per_prodi': per_prodi,
        'warna_prodi_list': warna_prodi_list,
        'legend_fakultas': legend_fakultas,
        'terbaru': terbaru, 'pengaturan': pengaturan,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/dashboard.html', context)


@login_required
def pendaftar(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    status    = request.GET.get('status', '')
    jalur     = request.GET.get('jalur', '')
    gelombang = request.GET.get('gelombang', '')
    cari      = request.GET.get('q', '')

    qs = Pendaftaran.objects.select_related(
        'user', 'jalur', 'prodi_pilihan_1', 'gelombang'
    ).order_by('-tgl_daftar')

    if status:    qs = qs.filter(status=status)
    if jalur:     qs = qs.filter(jalur_id=jalur)
    if gelombang: qs = qs.filter(gelombang_id=gelombang)
    if cari:
        qs = qs.filter(
            Q(user__first_name__icontains=cari) |
            Q(user__last_name__icontains=cari)  |
            Q(user__email__icontains=cari)       |
            Q(no_pendaftaran__icontains=cari)
        )

    # Ambil mapping kode_referral → nama recruiter
    from afiliasi.models import Recruiter
    recruiter_map = {
        r.kode_referral: r.user.get_full_name()
        for r in Recruiter.objects.select_related('user').all()
    }

    # Tambahkan nama recruiter ke setiap pendaftar
    pendaftar_list = list(qs)
    for p in pendaftar_list:
        p.nama_recruiter = recruiter_map.get(p.kode_referral, '')

    context = {
        'pendaftar_list':   pendaftar_list,
        'jalur_list':       JalurPenerimaan.objects.filter(status='aktif'),
        'gelombang_list':   GelombangPenerimaan.objects.all(),
        'status_choices':   Pendaftaran.STATUS_CHOICES,
        'filter_status':    status,
        'filter_jalur':     jalur,
        'filter_gelombang': gelombang,
        'filter_cari':      cari,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/pendaftar.html', context)


@login_required
def detail_pendaftar(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    pendaftaran = get_object_or_404(
        Pendaftaran.objects.select_related('user', 'jalur', 'gelombang',
        'prodi_pilihan_1', 'prodi_pilihan_2'), pk=pk
    )
    try:
        profil = pendaftaran.profil
    except:
        profil = None

    dokumen_list = DokumenPendaftar.objects.filter(
        pendaftaran=pendaftaran
    ).select_related('persyaratan')
    # Tambahkan setelah query dokumen_list
    recruiter = None
    if pendaftaran.kode_referral:
        from afiliasi.models import Recruiter
        try:
            recruiter = Recruiter.objects.select_related('user').get(
                kode_referral=pendaftaran.kode_referral
            )
        except Recruiter.DoesNotExist:
            pass
   
    log_list = pendaftaran.log_status.all().order_by('-tgl_perubahan')

    context = {
        'pendaftaran': pendaftaran,
        'profil': profil,
        'dokumen_list': dokumen_list,
        'log_list': log_list,
        'recruiter': recruiter,
        'status_choices': Pendaftaran.STATUS_CHOICES,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/detail_pendaftar.html', context)


@login_required
def ubah_status(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    if request.method == 'POST':
        pendaftaran = get_object_or_404(Pendaftaran, pk=pk)
        status_baru = request.POST.get('status_baru')
        keterangan  = request.POST.get('keterangan', '')

        from pendaftaran.models import LogStatusPendaftaran
        LogStatusPendaftaran.objects.create(
            pendaftaran  = pendaftaran,
            status_lama  = pendaftaran.status,
            status_baru  = status_baru,
            keterangan   = keterangan,
            diubah_oleh  = request.user,
        )
        pendaftaran.status = status_baru
        pendaftaran.save()

        # Kirim notifikasi otomatis
        trigger_map = {
            'LULUS_ADM':     'dokumen_acc',
            'TERJADWAL':     'terjadwal',
            'LULUS_SELEKSI': 'lulus_seleksi',
            'TIDAK_LULUS':   'tidak_lulus',
            'DAFTAR_ULANG':  'daftar_ulang',
        }
        trigger = trigger_map.get(status_baru)
        if trigger:
            try:
                from notifikasi.engine import kirim_notifikasi
                kirim_notifikasi(trigger, pendaftaran)
            except Exception as e:
                logger.error(f'Error notifikasi: {e}')

        messages.success(request, f'Status berhasil diubah ke {status_baru}.')

    return redirect('admin_pmb:detail_pendaftar', pk=pk)


@login_required
def verifikasi(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    dok_list = DokumenPendaftar.objects.filter(
        status_verifikasi='menunggu'
    ).select_related('pendaftaran__user', 'persyaratan').order_by('tgl_upload')

    context = {
        'dok_list': dok_list,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/verifikasi.html', context)


@login_required
def verif_acc(request, dok_id):
    if not cek_admin(request.user):
        return redirect('dashboard:index')
    if request.method == 'POST':
        dok = get_object_or_404(DokumenPendaftar, pk=dok_id)
        dok.status_verifikasi  = 'terverifikasi'
        dok.tgl_verifikasi     = timezone.now()
        dok.diverifikasi_oleh  = request.user
        dok.catatan_verifikasi = request.POST.get('catatan', '')
        dok.save()

        # Kirim notifikasi
        try:
            from notifikasi.engine import kirim_notifikasi
            kirim_notifikasi('dokumen_acc', dok.pendaftaran)
        except Exception as e:
            logger.error(f'Error notifikasi dokumen acc: {e}')

        messages.success(request, f'Dokumen "{dok.nama_dokumen}" berhasil diverifikasi.')
    return redirect('admin_pmb:verifikasi')


@login_required
def verif_tolak(request, dok_id):
    if not cek_admin(request.user):
        return redirect('dashboard:index')
    if request.method == 'POST':
        dok = get_object_or_404(DokumenPendaftar, pk=dok_id)
        dok.status_verifikasi  = 'ditolak'
        dok.tgl_verifikasi     = timezone.now()
        dok.diverifikasi_oleh  = request.user
        dok.catatan_verifikasi = request.POST.get('catatan', '')
        dok.save()
        messages.warning(request, f'Dokumen "{dok.nama_dokumen}" ditolak.')
    return redirect('admin_pmb:verifikasi')


@login_required
def pembayaran(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from django.core.paginator import Paginator
    from django.db.models import Count, Q, Sum
    from pembayaran.models import KonfirmasiPembayaran

    status = request.GET.get('status', 'menunggu')
    q = (request.GET.get('q') or '').strip()

    qs = KonfirmasiPembayaran.objects.select_related(
        'tagihan',
        'tagihan__pendaftaran',
        'tagihan__pendaftaran__user',
        'tagihan__pendaftaran__jalur',
        'rekening_tujuan',
        'dikonfirmasi_oleh',
    ).order_by('-created_at')

    if status in ('menunggu', 'dikonfirmasi', 'ditolak'):
        qs = qs.filter(status=status)

    if q:
        qs = qs.filter(
            Q(tagihan__kode_bayar__icontains=q)
            | Q(tagihan__pendaftaran__no_pendaftaran__icontains=q)
            | Q(atas_nama_pengirim__icontains=q)
            | Q(no_transaksi__icontains=q)
        )

    stats = KonfirmasiPembayaran.objects.aggregate(
        total=Count('id'),
        menunggu=Count('id', filter=Q(status='menunggu')),
        dikonfirmasi=Count('id', filter=Q(status='dikonfirmasi')),
        ditolak=Count('id', filter=Q(status='ditolak')),
        total_diterima=Sum('jumlah_bayar', filter=Q(status='dikonfirmasi')),
    )

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin_pmb/pembayaran.html', {
        'page_obj': page_obj,
        'status_active': status,
        'q': q,
        'stats': stats,
        'page_title': 'Verifikasi Pembayaran',
        **get_sidebar_counts()
    })

@login_required
def pembayaran_detail(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from django.contrib import messages
    from django.shortcuts import get_object_or_404
    from django.utils import timezone
    from pembayaran.models import KonfirmasiPembayaran

    konfirmasi = get_object_or_404(
        KonfirmasiPembayaran.objects.select_related(
            'tagihan',
            'tagihan__pendaftaran',
            'tagihan__pendaftaran__user',
            'tagihan__pendaftaran__jalur',
            'tagihan__pendaftaran__gelombang',
            'rekening_tujuan',
            'dikonfirmasi_oleh',
        ),
        pk=pk,
    )

    if request.method == 'POST':
        action = request.POST.get('action')
        catatan = (request.POST.get('catatan_admin') or '').strip()

        if konfirmasi.status != 'menunggu':
            messages.warning(request, "Konfirmasi ini sudah diproses sebelumnya.")
            return redirect('admin_pmb:pembayaran_detail', pk=pk)

        tagihan = konfirmasi.tagihan

        if action == 'approve':
            konfirmasi.status = 'dikonfirmasi'
            konfirmasi.catatan_admin = catatan
            konfirmasi.dikonfirmasi_oleh = request.user
            konfirmasi.tgl_konfirmasi = timezone.now()
            konfirmasi.save()

            tagihan.status = 'lunas'
            tagihan.save(update_fields=['status', 'updated_at'])

            KonfirmasiPembayaran.objects.filter(
                tagihan=tagihan, status='menunggu',
            ).exclude(pk=konfirmasi.pk).update(
                status='ditolak',
                catatan_admin='Auto-rejected: konfirmasi lain sudah disetujui.',
                dikonfirmasi_oleh=request.user,
                tgl_konfirmasi=timezone.now(),
            )

            messages.success(
                request,
                f"Pembayaran {tagihan.kode_bayar} berhasil dikonfirmasi sebagai LUNAS."
            )
            # TODO Step 6: trigger notifikasi 'konfirmasi_pembayaran'
            return redirect('admin_pmb:pembayaran')

        elif action == 'reject':
            if not catatan:
                messages.error(request, "Catatan alasan penolakan wajib diisi.")
                return redirect('admin_pmb:pembayaran_detail', pk=pk)

            konfirmasi.status = 'ditolak'
            konfirmasi.catatan_admin = catatan
            konfirmasi.dikonfirmasi_oleh = request.user
            konfirmasi.tgl_konfirmasi = timezone.now()
            konfirmasi.save()

            has_other_active = tagihan.konfirmasi.filter(
                status__in=('menunggu', 'dikonfirmasi')
            ).exists()
            if not has_other_active:
                tagihan.status = 'belum_bayar'
                tagihan.save(update_fields=['status', 'updated_at'])

            messages.success(
                request,
                "Konfirmasi pembayaran ditolak. Maba bisa upload ulang bukti."
            )
            return redirect('admin_pmb:pembayaran')

    return render(request, 'admin_pmb/pembayaran_detail.html', {
        'konfirmasi': konfirmasi,
        'tagihan': konfirmasi.tagihan,
        'page_title': f'Verifikasi {konfirmasi.tagihan.kode_bayar}',
        **get_sidebar_counts()
    })

@login_required
def seleksi(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from seleksi.models import JadwalSeleksi
    from master.models import JalurPenerimaan, GelombangPenerimaan

    jadwal_list = JadwalSeleksi.objects.select_related(
        'jalur', 'gelombang', 'dibuat_oleh'
    ).order_by('-tgl_seleksi')

    context = {
        'jadwal_list':    jadwal_list,
        'jalur_list':     JalurPenerimaan.objects.filter(status='aktif'),
        'gelombang_list': GelombangPenerimaan.objects.all(),
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/seleksi.html', context)

@login_required
def hasil(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from seleksi.models import HasilPenerimaan
    from pendaftaran.models import Pendaftaran
    from master.models import ProdiPMB, GelombangPenerimaan

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'input_hasil':
            pendaftaran_id = request.POST.get('pendaftaran_id')
            status_hasil   = request.POST.get('status_hasil')
            prodi_id       = request.POST.get('prodi_id')
            nilai_akhir    = request.POST.get('nilai_akhir') or None
            peringkat      = request.POST.get('peringkat') or None
            catatan        = request.POST.get('catatan', '')

            p = get_object_or_404(Pendaftaran, pk=pendaftaran_id)
            HasilPenerimaan.objects.update_or_create(
                pendaftaran=p,
                defaults={
                    'prodi_diterima_id': prodi_id,
                    'status':            status_hasil,
                    'nilai_akhir':       nilai_akhir,
                    'peringkat':         peringkat,
                    'catatan':           catatan,
                    'diinput_oleh':      request.user,
                }
            )
            if status_hasil == 'lulus':
                p.status = 'LULUS_SELEKSI'
            elif status_hasil == 'tidak_lulus':
                p.status = 'TIDAK_LULUS'
            p.save()
            messages.success(request, 'Hasil penerimaan berhasil disimpan.')
            return redirect('admin_pmb:hasil')

    gelombang_id = request.GET.get('gelombang', '')
    prodi_id     = request.GET.get('prodi', '')
    status_f     = request.GET.get('status', '')

    qs = Pendaftaran.objects.filter(
        status__in=['TERJADWAL', 'LULUS_ADM', 'LULUS_SELEKSI', 'TIDAK_LULUS']
    ).select_related('user', 'prodi_pilihan_1', 'gelombang', 'hasil')

    if gelombang_id: qs = qs.filter(gelombang_id=gelombang_id)
    if prodi_id:     qs = qs.filter(prodi_pilihan_1_id=prodi_id)
    if status_f:     qs = qs.filter(status=status_f)

    context = {
        'pendaftar_list':   qs,
        'prodi_list':       ProdiPMB.objects.filter(status='aktif'),
        'gelombang_list':   GelombangPenerimaan.objects.all(),
        'filter_gelombang': gelombang_id,
        'filter_prodi':     prodi_id,
        'filter_status':    status_f,
        'total_lulus':      Pendaftaran.objects.filter(status='LULUS_SELEKSI').count(),
        'total_tidak_lulus':Pendaftaran.objects.filter(status='TIDAK_LULUS').count(),
        'total_terjadwal':  Pendaftaran.objects.filter(status='TERJADWAL').count(),
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/hasil.html', context)

@login_required
def master(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from master.models import (JalurPenerimaan, GelombangPenerimaan,
                                ProdiPMB, PengaturanSistem)

    menu_master = [
        {'nama': 'Jalur Penerimaan',   'icon': '🚪', 'url': '/admin/master/jalurpenerimaan/',    'desc': 'Kelola jalur masuk'},
        {'nama': 'Gelombang',          'icon': '🌊', 'url': '/admin/master/gelombangpenerimaan/','desc': 'Jadwal & biaya gelombang'},
        {'nama': 'Program Studi',      'icon': '🎓', 'url': '/admin/master/prodipmb/',           'desc': 'Prodi yang dibuka PMB'},
        {'nama': 'Persyaratan Jalur',  'icon': '📋', 'url': '/admin/master/persyaratanjalur/',   'desc': 'Dokumen per jalur'},
        {'nama': 'Pengaturan Sistem',  'icon': '⚙️', 'url': '/admin/master/pengaturansistem/',  'desc': 'Konfigurasi umum PMB'},
        {'nama': 'Pengaturan Afiliasi','icon': '💰', 'url': '/admin/afiliasi/pengaturanafiliasi/','desc': 'Nominal komisi recruiter'},
    ]

    context = {
        'menu_master':    menu_master,
        'jalur_list':     JalurPenerimaan.objects.all().order_by('urutan'),
        'gelombang_list': GelombangPenerimaan.objects.all().order_by('-tgl_buka'),
        'prodi_list':     ProdiPMB.objects.filter(status='aktif').order_by('nama_prodi'),
        'pengaturan':     PengaturanSistem.get(),
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/master.html', context)

@login_required
def konten(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from konten.models import (Pengumuman, Testimoni, MitraKerjasama,
                                FAQ, BrosurFakultas, DokumenDownload)

    menu_konten = [
        {'nama': 'Pengumuman',      'icon': '📢', 'url': '/admin/konten/pengumuman/',    'desc': 'Kelola pengumuman PMB'},
        {'nama': 'Testimoni',       'icon': '💬', 'url': '/admin/konten/testimoni/',     'desc': 'Testimoni mahasiswa & alumni'},
        {'nama': 'Mitra Kerjasama', 'icon': '🤝', 'url': '/admin/konten/mitrakerjasama/','desc': 'Logo & link mitra kampus'},
        {'nama': 'FAQ',             'icon': '❓', 'url': '/admin/konten/faq/',           'desc': 'Pertanyaan yang sering ditanyakan'},
        {'nama': 'Brosur Fakultas', 'icon': '📄', 'url': '/admin/konten/brosurfakultas/','desc': 'Upload brosur per fakultas'},
        {'nama': 'Dokumen Download','icon': '📥', 'url': '/admin/konten/dokumendownload/','desc': 'Formulir & dokumen unduhan'},
        {'nama': 'Galeri Kampus',   'icon': '🖼️', 'url': '/admin/konten/galerikampus/', 'desc': 'Foto & video kampus'},
        {'nama': 'Media Sosial',    'icon': '📱', 'url': '/admin/konten/mediasosial/',  'desc': 'Link media sosial UNISAN'},
    ]

    context = {
        'menu_konten':      menu_konten,
        'total_pengumuman': Pengumuman.objects.filter(status='aktif').count(),
        'total_testimoni':  Testimoni.objects.filter(status='aktif').count(),
        'total_mitra':      MitraKerjasama.objects.filter(status='aktif').count(),
        'total_faq':        FAQ.objects.filter(status='aktif').count(),
        'total_brosur':     BrosurFakultas.objects.filter(status='aktif').count(),
        'total_dokumen':    DokumenDownload.objects.filter(status='aktif').count(),
        'pengumuman_list':  Pengumuman.objects.order_by('-tgl_tayang')[:5],
        'testimoni_list':   Testimoni.objects.order_by('-id')[:5],
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/konten.html', context)

@login_required
def afiliasi(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from afiliasi.models import Recruiter, KomisiReferral, PencairanKomisi
    from django.db.models import Sum

    status_filter = request.GET.get('status', 'menunggu')
    status_tabs = [
        ('menunggu',  'Menunggu Verifikasi', '#f59e0b'),
        ('aktif',     'Aktif',               '#10b981'),
        ('nonaktif',  'Ditolak',             '#ef4444'),
        ('suspend',   'Suspend',             '#94a3b8'),
    ]

    recruiter_list = Recruiter.objects.filter(
        status=status_filter
    ).select_related('user').order_by('-tgl_bergabung')

    total_menunggu  = Recruiter.objects.filter(status='menunggu').count()
    total_aktif     = Recruiter.objects.filter(status='aktif').count()
    total_komisi    = KomisiReferral.objects.filter(status='approved').aggregate(
        total=Sum('jumlah_komisi'))['total'] or 0
    total_pencairan = PencairanKomisi.objects.filter(status='pending').count()

    context = {
        'recruiter_list':  recruiter_list,
        'status_filter':   status_filter,
        'status_tabs':     status_tabs,
        'total_menunggu':  total_menunggu,
        'total_aktif':     total_aktif,
        'total_komisi':    total_komisi,
        'total_pencairan': total_pencairan,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/afiliasi.html', context)


@login_required
def laporan(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from pendaftaran.models import Pendaftaran, ProfilPendaftar
    from django.db.models import Count, Q

    # Statistik umum
    total_pendaftar = Pendaftaran.objects.count()
    total_lulus     = Pendaftaran.objects.filter(status='LULUS_SELEKSI').count()
    total_daftar_ulang = Pendaftaran.objects.filter(status='DAFTAR_ULANG').count()

    # Per prodi (+ warna per fakultas)
    from accounts.views import warna_fakultas, WARNA_FAKULTAS
    per_prodi_qs = Pendaftaran.objects.values(
        'prodi_pilihan_1__nama_prodi',
        'prodi_pilihan_1__kode_fakultas',
    ).annotate(total=Count('id')).order_by('-total')

    per_prodi = []
    warna_prodi_list = []
    for p in per_prodi_qs:
        kode_fak = p['prodi_pilihan_1__kode_fakultas']
        w = warna_fakultas(kode_fak)
        per_prodi.append({
            'prodi_pilihan_1__nama_prodi': p['prodi_pilihan_1__nama_prodi'],
            'total': p['total'],
            'kode_fakultas': kode_fak,
            'warna': w['text'],
        })
        warna_prodi_list.append(w['text'])

    # Legend fakultas — hanya yang muncul di chart
    kode_fakultas_muncul = {p['kode_fakultas'] for p in per_prodi if p['kode_fakultas']}
    legend_fakultas = [
        {'kode': k, 'hex': WARNA_FAKULTAS[k]['text']}
        for k in kode_fakultas_muncul if k in WARNA_FAKULTAS
    ]

    # Per jalur
    per_jalur = Pendaftaran.objects.values(
        'jalur__nama_jalur'
    ).annotate(total=Count('id')).order_by('-total')

    # Per gelombang
    per_gelombang = Pendaftaran.objects.values(
        'gelombang__nama_gelombang'
    ).annotate(total=Count('id')).order_by('-total')

    # Per provinsi
    per_provinsi = ProfilPendaftar.objects.exclude(
        provinsi_nama=''
    ).values('provinsi_nama').annotate(
        total=Count('id')
    ).order_by('-total')[:15]

    # Per kab/kota
    per_kabkota = ProfilPendaftar.objects.exclude(
        kabupaten_kota_nama=''
    ).values('kabupaten_kota_nama', 'provinsi_nama').annotate(
        total=Count('id')
    ).order_by('-total')[:15]

    # Per asal sekolah
    per_sekolah = ProfilPendaftar.objects.exclude(
        asal_sekolah=''
    ).values('asal_sekolah', 'kabupaten_kota_nama').annotate(
        total=Count('id')
    ).order_by('-total')[:20]

 # Per sumber informasi (multi-select JSONField)
    per_sumber = ProfilPendaftar.objects.exclude(
        sumber_informasi=[]
    ).exclude(
        sumber_informasi__isnull=True
    ).values('sumber_informasi').annotate(
        total=Count('id')
    ).order_by('-total')


    # Ukuran baju
    per_ukuran = ProfilPendaftar.objects.exclude(
        ukuran_baju=''
    ).values('ukuran_baju').annotate(
        total=Count('id')
    ).order_by('ukuran_baju')

    # Per jenis kelamin
    per_gender = ProfilPendaftar.objects.values(
        'jenis_kelamin'
    ).annotate(total=Count('id')).order_by('jenis_kelamin')

    # Per agama
    per_agama = ProfilPendaftar.objects.exclude(
        agama_nama=''
    ).values('agama_nama').annotate(
        total=Count('id')
    ).order_by('-total')

    # Tren per bulan
    from django.db.models.functions import TruncMonth
    per_bulan = Pendaftaran.objects.annotate(
        bulan=TruncMonth('tgl_daftar')
    ).values('bulan').annotate(
        total=Count('id')
    ).order_by('bulan')

    tabs = [
        {'id': 'prodi',   'label': '📊 Prodi & Jalur'},
        {'id': 'wilayah', 'label': '🗺️ Wilayah'},
        {'id': 'sekolah', 'label': '🏫 Sekolah Asal'},
        {'id': 'sumber',  'label': '📣 Sumber Promosi'},
        {'id': 'baju',    'label': '👕 Ukuran Baju'},
        {'id': 'demo',    'label': '👥 Demografi'},
    ]

    context = {
        'total_pendaftar':    total_pendaftar,
        'total_lulus':        total_lulus,
        'total_daftar_ulang': total_daftar_ulang,
        'per_prodi':          per_prodi,
        'per_jalur':          per_jalur,
        'per_gelombang':      per_gelombang,
        'per_provinsi':       per_provinsi,
        'per_kabkota':        per_kabkota,
        'per_sekolah':        per_sekolah,
        'per_sumber':         per_sumber,
        'per_ukuran':         per_ukuran,
        'per_gender':         per_gender,
        'per_agama':          per_agama,
        'per_bulan':          per_bulan,
        'tabs':               tabs,
        'warna_prodi_list':  warna_prodi_list,
        'legend_fakultas':   legend_fakultas,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/laporan.html', context)

@login_required
def chatbot_kb(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')
    kb_list = KnowledgeBase.objects.all().order_by('kategori', 'urutan_prioritas')
    return render(request, 'admin_pmb/chatbot_kb.html', {
        'kb_list': kb_list,
        **get_sidebar_counts()
    })

@login_required
def afiliasi(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from afiliasi.models import Recruiter, KomisiReferral, PencairanKomisi
    from django.db.models import Sum

    status_filter = request.GET.get('status', 'menunggu')

    recruiter_list = Recruiter.objects.filter(
        status=status_filter
    ).select_related('user').order_by('-tgl_bergabung')

    # Statistik
    total_menunggu  = Recruiter.objects.filter(status='menunggu').count()
    total_aktif     = Recruiter.objects.filter(status='aktif').count()
    total_komisi    = KomisiReferral.objects.filter(status='approved').aggregate(
        total=Sum('jumlah_komisi'))['total'] or 0
    total_pencairan = PencairanKomisi.objects.filter(status='pending').count()

    context = {
        'recruiter_list':  recruiter_list,
        'status_filter':   status_filter,
        'total_menunggu':  total_menunggu,
        'total_aktif':     total_aktif,
        'total_komisi':    total_komisi,
        'total_pencairan': total_pencairan,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/afiliasi.html', context)


@login_required
def afiliasi_approve(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')
    if request.method == 'POST':
        from afiliasi.models import Recruiter
        rec = get_object_or_404(Recruiter, pk=pk)
        rec.status = 'aktif'
        rec.save()
        messages.success(request, f'Recruiter {rec.user.get_full_name()} berhasil diaktifkan!')
    return redirect('admin_pmb:afiliasi')


@login_required
def afiliasi_tolak(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')
    if request.method == 'POST':
        from afiliasi.models import Recruiter
        rec = get_object_or_404(Recruiter, pk=pk)
        alasan = request.POST.get('alasan', '')
        rec.status  = 'nonaktif'
        rec.catatan = f'DITOLAK: {alasan}'
        rec.save()
        messages.warning(request, f'Recruiter {rec.user.get_full_name()} ditolak.')
    return redirect('admin_pmb:afiliasi')


@login_required
def afiliasi_detail(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')
    from afiliasi.models import Recruiter, KomisiReferral
    rec = get_object_or_404(Recruiter.objects.select_related('user'), pk=pk)
    komisi_list = KomisiReferral.objects.filter(
        recruiter=rec).select_related('pendaftaran__user').order_by('-tgl_komisi')
    context = {
        'rec': rec,
        'komisi_list': komisi_list,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/afiliasi_detail.html', context)

@login_required
def seleksi_tambah(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from seleksi.models import JadwalSeleksi
    from master.models import JalurPenerimaan, GelombangPenerimaan

    if request.method == 'POST':
        JadwalSeleksi.objects.create(
            jalur         = JalurPenerimaan.objects.get(pk=request.POST['jalur']),
            gelombang     = GelombangPenerimaan.objects.get(pk=request.POST['gelombang']),
            jenis_seleksi = request.POST['jenis_seleksi'],
            nama_seleksi  = request.POST['nama_seleksi'],
            tgl_seleksi   = request.POST['tgl_seleksi'],
            jam_mulai     = request.POST['jam_mulai'],
            jam_selesai   = request.POST['jam_selesai'],
            lokasi        = request.POST.get('lokasi', ''),
            link_online   = request.POST.get('link_online', ''),
            keterangan    = request.POST.get('keterangan', ''),
            status        = 'draft',
            dibuat_oleh   = request.user,
        )
        messages.success(request, 'Jadwal seleksi berhasil ditambahkan!')
        return redirect('admin_pmb:seleksi')

    from seleksi.models import JadwalSeleksi as JS
    context = {
        'jalur_list':     JalurPenerimaan.objects.filter(status='aktif'),
        'gelombang_list': GelombangPenerimaan.objects.all(),
        'jenis_choices':  JS.JENIS_CHOICES,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/seleksi_tambah.html', context)


@login_required
def seleksi_detail(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from seleksi.models import JadwalSeleksi, PesertaSeleksi
    from pendaftaran.models import Pendaftaran

    jadwal = get_object_or_404(JadwalSeleksi, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'publish':
            pendaftar = Pendaftaran.objects.filter(
                jalur=jadwal.jalur,
                gelombang=jadwal.gelombang,
                status__in=['LENGKAP', 'LULUS_ADM', 'TERJADWAL']
            )
            added = 0
            for p in pendaftar:
                obj, created = PesertaSeleksi.objects.get_or_create(
                    jadwal=jadwal, pendaftaran=p,
                    defaults={'no_ujian': f'{jadwal.pk:03d}{p.pk:04d}'}
                )
                if created:
                    added += 1
                    p.status = 'TERJADWAL'
                    p.save()
            jadwal.status = 'publish'
            jadwal.save()
            messages.success(request, f'Jadwal dipublikasi! {added} peserta ditambahkan.')

        elif action == 'nilai':
            peserta_id = request.POST.get('peserta_id')
            nilai      = request.POST.get('nilai')
            hadir      = request.POST.get('hadir')
            catatan    = request.POST.get('catatan', '')
            from django.utils import timezone
            PesertaSeleksi.objects.filter(pk=peserta_id).update(
                nilai            = nilai if nilai else None,
                status_kehadiran = hadir,
                catatan_penilai  = catatan,
                dinilai_oleh     = request.user,
                tgl_penilaian    = timezone.now(),
            )
            messages.success(request, 'Nilai berhasil disimpan.')

        elif action == 'selesai':
            jadwal.status = 'selesai'
            jadwal.save()
            messages.success(request, 'Jadwal seleksi ditandai selesai.')

        return redirect('admin_pmb:seleksi_detail', pk=pk)

    peserta_list = jadwal.peserta.select_related(
        'pendaftaran__user', 'pendaftaran__prodi_pilihan_1'
    ).order_by('no_ujian')

    context = {
        'jadwal':       jadwal,
        'peserta_list': peserta_list,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/seleksi_detail.html', context)

@login_required
def export_pendaftar(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from pendaftaran.models import Pendaftaran, ProfilPendaftar

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Pendaftar'

    # Style
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(fill_type='solid', fgColor='667EEA')
    center       = Alignment(horizontal='center', vertical='center')
    thin         = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Header
    headers = [
        'No', 'No. Pendaftaran', 'Nama Lengkap', 'Email', 'No HP',
        'Jalur', 'Gelombang', 'Prodi Pilihan 1', 'Prodi Pilihan 2',
        'Status', 'Tgl Daftar',
        'NIK', 'Tempat Lahir', 'Tgl Lahir', 'Jenis Kelamin', 'Agama',
        'Alamat', 'Provinsi', 'Kab/Kota', 'Kecamatan', 'Kode Pos',
        'Nama Ayah', 'Pekerjaan Ayah', 'Nama Ibu', 'Pekerjaan Ibu', 'No HP Ortu',
        'Asal Sekolah', 'Jurusan Sekolah', 'Tahun Lulus', 'Nilai Rata-rata',
        'Ukuran Baju', 'Sumber Informasi', 'Kode Referral',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin

    # Lebar kolom
    col_widths = [5,20,25,30,15,12,15,25,25,15,15,
                  18,15,12,12,12,35,20,20,20,10,
                  25,20,25,20,15,
                  35,20,12,8,
                  10,20,15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    # Freeze header
    ws.freeze_panes = 'A2'

    # Data
    pendaftar_list = Pendaftaran.objects.select_related(
        'user', 'jalur', 'gelombang', 'prodi_pilihan_1', 'prodi_pilihan_2'
    ).order_by('tgl_daftar')

    for row_num, p in enumerate(pendaftar_list, 2):
        try:
            profil = p.profil
        except:
            profil = None

        data = [
            row_num - 1,
            p.no_pendaftaran,
            p.user.get_full_name(),
            p.user.email,
            p.user.no_hp,
            p.jalur.nama_jalur,
            p.gelombang.nama_gelombang,
            p.prodi_pilihan_1.nama_prodi if p.prodi_pilihan_1 else '',
            p.prodi_pilihan_2.nama_prodi if p.prodi_pilihan_2 else '',
            p.get_status_display(),
            p.tgl_daftar.strftime('%d/%m/%Y') if p.tgl_daftar else '',
            profil.nik if profil else '',
            profil.tempat_lahir if profil else '',
            profil.tgl_lahir.strftime('%d/%m/%Y') if profil and profil.tgl_lahir else '',
            profil.get_jenis_kelamin_display() if profil and profil.jenis_kelamin else '',
            profil.agama_nama if profil else '',
            profil.alamat_lengkap if profil else '',
            profil.provinsi_nama if profil else '',
            profil.kabupaten_kota_nama if profil else '',
            profil.kecamatan_nama if profil else '',
            profil.kode_pos if profil else '',
            profil.nama_ayah if profil else '',
            profil.pekerjaan_ayah if profil else '',
            profil.nama_ibu if profil else '',
            profil.pekerjaan_ibu if profil else '',
            profil.no_hp_ortu if profil else '',
            profil.asal_sekolah if profil else '',
            profil.jurusan_sekolah if profil else '',
            profil.tahun_lulus if profil else '',
            str(profil.nilai_rata_rata) if profil and profil.nilai_rata_rata else '',
            profil.ukuran_baju if profil else '',
            ', '.join([dict(ProfilPendaftar.SUMBER_INFO_CHOICES).get(s, s) for s in (profil.sumber_informasi or [])]) if profil else '',
           
            p.kode_referral or '',
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border    = thin
            cell.alignment = Alignment(vertical='center')
            if row_num % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='F8F9FF')

    # Auto filter
    ws.auto_filter.ref = f'A1:{ws.cell(1, len(headers)).column_letter}1'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="data_pendaftar_pmb.xlsx"'
    wb.save(response)
    return response


@login_required
def export_ukuran_baju(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from django.db.models import Count
    from pendaftaran.models import ProfilPendaftar, Pendaftaran

    wb = openpyxl.Workbook()

    # Sheet 1 — Detail per pendaftar
    ws1 = wb.active
    ws1.title = 'Detail Ukuran Baju'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(fill_type='solid', fgColor='667EEA')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers1 = ['No', 'No. Pendaftaran', 'Nama', 'Prodi', 'Jalur', 'Ukuran Baju']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 28
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 12
    ws1.freeze_panes = 'A2'

    profil_list = ProfilPendaftar.objects.exclude(
        ukuran_baju=''
    ).select_related(
        'pendaftaran__user',
        'pendaftaran__prodi_pilihan_1',
        'pendaftaran__jalur'
    ).order_by('ukuran_baju', 'pendaftaran__prodi_pilihan_1__nama_prodi')

    for i, pr in enumerate(profil_list, 2):
        data = [
            i - 1,
            pr.pendaftaran.no_pendaftaran,
            pr.pendaftaran.user.get_full_name(),
            pr.pendaftaran.prodi_pilihan_1.nama_prodi if pr.pendaftaran.prodi_pilihan_1 else '',
            pr.pendaftaran.jalur.nama_jalur,
            pr.ukuran_baju,
        ]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.border = thin
            if i % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='F8F9FF')

    # Sheet 2 — Rekap per ukuran
    ws2 = wb.create_sheet('Rekap Ukuran')
    headers2 = ['Ukuran', 'Jumlah', 'Persentase']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 15

    rekap = ProfilPendaftar.objects.exclude(
        ukuran_baju=''
    ).values('ukuran_baju').annotate(
        total=Count('id')
    ).order_by('ukuran_baju')

    total_all = sum(r['total'] for r in rekap)

    for i, r in enumerate(rekap, 2):
        pct = f"{(r['total']/total_all*100):.1f}%" if total_all else '0%'
        for col, val in enumerate([r['ukuran_baju'], r['total'], pct], 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center')

    # Total
    total_row = len(list(rekap)) + 2
    ws2.cell(total_row, 1, 'TOTAL').font = Font(bold=True)
    ws2.cell(total_row, 2, total_all).font = Font(bold=True)
    ws2.cell(total_row, 3, '100%').font = Font(bold=True)

    # Sheet 3 — Rekap per prodi per ukuran
    ws3 = wb.create_sheet('Per Prodi')
    ws3.cell(1, 1, 'Prodi').font = header_font
    ws3.cell(1, 1).fill = header_fill

    ukuran_list = ['S', 'M', 'L', 'XL', 'XXL', '3XL']
    for col, uk in enumerate(ukuran_list, 2):
        cell = ws3.cell(1, col, uk)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    ws3.cell(1, len(ukuran_list)+2, 'Total').font = header_font
    ws3.cell(1, len(ukuran_list)+2).fill = header_fill

    from master.models import ProdiPMB
    prodi_list = ProdiPMB.objects.filter(status='aktif').order_by('nama_prodi')

    for row, prodi in enumerate(prodi_list, 2):
        ws3.cell(row, 1, prodi.nama_prodi)
        total_prodi = 0
        for col, uk in enumerate(ukuran_list, 2):
            jumlah = ProfilPendaftar.objects.filter(
                ukuran_baju=uk,
                pendaftaran__prodi_pilihan_1=prodi
            ).count()
            ws3.cell(row, col, jumlah).alignment = Alignment(horizontal='center')
            total_prodi += jumlah
        ws3.cell(row, len(ukuran_list)+2, total_prodi).font = Font(bold=True)

    ws3.column_dimensions['A'].width = 30

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rekap_ukuran_baju.xlsx"'
    wb.save(response)
    return response


@login_required
def export_wilayah(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from django.db.models import Count
    from pendaftaran.models import ProfilPendaftar

    wb  = openpyxl.Workbook()
    hf  = Font(bold=True, color='FFFFFF')
    hfi = PatternFill(fill_type='solid', fgColor='667EEA')
    th  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def set_header(ws, headers, widths):
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            c.font = hf; c.fill = hfi
            c.alignment = Alignment(horizontal='center')
            c.border = th
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1,i).column_letter].width = w
        ws.freeze_panes = 'A2'

    # Sheet 1 — Per Provinsi
    ws1 = wb.active
    ws1.title = 'Per Provinsi'
    set_header(ws1, ['No','Provinsi','Jumlah Pendaftar','%'], [5,30,20,10])
    per_prov = ProfilPendaftar.objects.exclude(provinsi_nama='').values(
        'provinsi_nama').annotate(total=Count('id')).order_by('-total')
    total = sum(p['total'] for p in per_prov)
    for i, p in enumerate(per_prov, 2):
        pct = f"{p['total']/total*100:.1f}%" if total else '0%'
        for col, val in enumerate([i-1, p['provinsi_nama'], p['total'], pct], 1):
            c = ws1.cell(i, col, val); c.border = th
            if i%2==0: c.fill = PatternFill(fill_type='solid', fgColor='F8F9FF')

    # Sheet 2 — Per Kab/Kota
    ws2 = wb.create_sheet('Per Kab-Kota')
    set_header(ws2, ['No','Provinsi','Kab/Kota','Jumlah','%'], [5,25,25,15,10])
    per_kab = ProfilPendaftar.objects.exclude(kabupaten_kota_nama='').values(
        'provinsi_nama','kabupaten_kota_nama').annotate(
        total=Count('id')).order_by('-total')
    total2 = sum(p['total'] for p in per_kab)
    for i, p in enumerate(per_kab, 2):
        pct = f"{p['total']/total2*100:.1f}%" if total2 else '0%'
        for col, val in enumerate([i-1, p['provinsi_nama'],
                                   p['kabupaten_kota_nama'], p['total'], pct], 1):
            c = ws2.cell(i, col, val); c.border = th
            if i%2==0: c.fill = PatternFill(fill_type='solid', fgColor='F8F9FF')

    # Sheet 3 — Per Sekolah
    ws3 = wb.create_sheet('Per Sekolah')
    set_header(ws3, ['No','Asal Sekolah','Kab/Kota','Jumlah Pendaftar'], [5,40,25,18])
    per_skl = ProfilPendaftar.objects.exclude(asal_sekolah='').values(
        'asal_sekolah','kabupaten_kota_nama').annotate(
        total=Count('id')).order_by('-total')
    for i, p in enumerate(per_skl, 2):
        for col, val in enumerate([i-1, p['asal_sekolah'],
                                   p['kabupaten_kota_nama'], p['total']], 1):
            c = ws3.cell(i, col, val); c.border = th
            if i%2==0: c.fill = PatternFill(fill_type='solid', fgColor='F8F9FF')

    # Sheet 4 — Sumber Informasi
    ws4 = wb.create_sheet('Sumber Informasi')
    set_header(ws4, ['No','Sumber Informasi','Jumlah','%'], [5,30,15,10])

    # Per sumber informasi (multi-select JSONField)
    from collections import Counter
    all_sumber = []
    for profil in ProfilPendaftar.objects.exclude(sumber_informasi=[]).exclude(sumber_informasi__isnull=True):
        if isinstance(profil.sumber_informasi, list):
            all_sumber.extend(profil.sumber_informasi)
    counter = Counter(all_sumber)
    total4 = sum(counter.values())
    sumber_map = dict(ProfilPendaftar.SUMBER_INFO_CHOICES)
    for i, (kode, total) in enumerate(counter.most_common(), 2):
        pct = f"{total/total4*100:.1f}%" if total4 else '0%'
        label = sumber_map.get(kode, kode)
        for col, val in enumerate([i-1, label, total, pct], 1):
            c = ws4.cell(i, col, val); c.border = th
            if i%2==0: c.fill = PatternFill(fill_type='solid', fgColor='F8F9FF')


    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rekap_wilayah_pmb.xlsx"'
    wb.save(response)
    return response

@login_required
def notifikasi(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from notifikasi.models import TemplateNotifikasi, LogNotifikasi

    template_list = TemplateNotifikasi.objects.all().order_by('trigger')
    log_terbaru   = LogNotifikasi.objects.select_related(
        'user', 'template'
    ).order_by('-tgl_kirim')[:20]

    context = {
        'template_list': template_list,
        'log_terbaru':   log_terbaru,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/notifikasi.html', context)


@login_required
def notifikasi_kirim(request):
    """Kirim notifikasi massal manual"""
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    if request.method == 'POST':
        status_target = request.POST.get('status_target', '')
        subjek        = request.POST.get('subjek', '')
        isi_email     = request.POST.get('isi_email', '')
        isi_wa        = request.POST.get('isi_wa', '')
        kirim_ke      = request.POST.get('kirim_ke', 'both')

        qs = Pendaftaran.objects.select_related('user', 'jalur', 'prodi_pilihan_1')
        if status_target:
            qs = qs.filter(status=status_target)

        from notifikasi.engine import kirim_notifikasi_manual
        berhasil, gagal = kirim_notifikasi_manual(
            list(qs), subjek, isi_email, isi_wa, kirim_ke
        )
        messages.success(
            request,
            f'Notifikasi terkirim: {berhasil} berhasil, {gagal} gagal dari {qs.count()} target.'
        )

    return redirect('admin_pmb:notifikasi')



@login_required
def notifikasi(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from notifikasi.models import TemplateNotifikasi, LogNotifikasi

    template_list = TemplateNotifikasi.objects.all().order_by('trigger')
    log_terbaru   = LogNotifikasi.objects.select_related(
        'user', 'template'
    ).order_by('-tgl_kirim')[:20]

    variabel_list = [
        'nama', 'no_pendaftaran', 'jalur', 'gelombang',
        'prodi', 'status', 'email', 'no_hp',
        'url_dashboard', 'nama_kampus', 'kontak_pmb',
    ]

    context = {
        'template_list':  template_list,
        'log_terbaru':    log_terbaru,
        'variabel_list':  variabel_list,
        'status_choices': Pendaftaran.STATUS_CHOICES,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/notifikasi.html', context)

@login_required
def notifikasi_log(request):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from notifikasi.models import LogNotifikasi
    from django.db.models import Count
    from django.utils.dateparse import parse_date

    # Handle hapus
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'hapus':
            hapus_filter = request.POST.get('hapus_filter', '0')

            if hapus_filter == '1':
                # Hapus berdasarkan filter
                qs = LogNotifikasi.objects.all()
                dari   = request.POST.get('filter_dari', '')
                sampai = request.POST.get('filter_sampai', '')
                status = request.POST.get('filter_status', '')
                jenis  = request.POST.get('filter_jenis', '')
                if dari:    qs = qs.filter(tgl_kirim__date__gte=dari)
                if sampai:  qs = qs.filter(tgl_kirim__date__lte=sampai)
                if status:  qs = qs.filter(status=status)
                if jenis:   qs = qs.filter(jenis=jenis)
                jumlah = qs.count()
                qs.delete()
                messages.success(request, f'{jumlah} log berhasil dihapus.')
            else:
                # Hapus yang dicentang
                log_ids = request.POST.getlist('log_ids')
                if log_ids:
                    jumlah = LogNotifikasi.objects.filter(pk__in=log_ids).count()
                    LogNotifikasi.objects.filter(pk__in=log_ids).delete()
                    messages.success(request, f'{jumlah} log berhasil dihapus.')

        return redirect('admin_pmb:notifikasi_log')

    # Filter
    filter_dari   = request.GET.get('dari', '')
    filter_sampai = request.GET.get('sampai', '')
    filter_status = request.GET.get('status', '')
    filter_jenis  = request.GET.get('jenis', '')

    qs = LogNotifikasi.objects.select_related(
        'user', 'template', 'pendaftaran'
    ).order_by('-tgl_kirim')

    if filter_dari:   qs = qs.filter(tgl_kirim__date__gte=filter_dari)
    if filter_sampai: qs = qs.filter(tgl_kirim__date__lte=filter_sampai)
    if filter_status: qs = qs.filter(status=filter_status)
    if filter_jenis:  qs = qs.filter(jenis=filter_jenis)

    stat = LogNotifikasi.objects.values('status').annotate(total=Count('id'))

    context = {
        'log_list':      qs[:200],
        'stat':          stat,
        'filter_dari':   filter_dari,
        'filter_sampai': filter_sampai,
        'filter_status': filter_status,
        'filter_jenis':  filter_jenis,
        **get_sidebar_counts(),
    }
    return render(request, 'admin_pmb/notifikasi_log.html', context)

@login_required
def cetak_kartu(request, pk):
    """Cetak kartu peserta satu pendaftar"""
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from django.http import HttpResponse
    from seleksi.kartu_pdf import buat_kartu_peserta
    from seleksi.models import KartuPeserta

    pendaftaran = get_object_or_404(Pendaftaran, pk=pk)

    # Generate atau ambil nomor kartu
    kartu, created = KartuPeserta.objects.get_or_create(
        pendaftaran=pendaftaran,
        defaults={
            'no_kartu': f'PMB-{pendaftaran.gelombang.kode_gelombang}-{pendaftaran.pk:04d}'
                        if hasattr(pendaftaran.gelombang, 'kode_gelombang')
                        else f'PMB-{pendaftaran.no_pendaftaran}',
        }
    )
    kartu.sudah_cetak = True
    kartu.save()

    # Ambil jadwal terbaru jika ada
    jadwal = None
    try:
        from seleksi.models import PesertaSeleksi
        peserta = PesertaSeleksi.objects.filter(
            pendaftaran=pendaftaran
        ).select_related('jadwal').last()
        if peserta:
            jadwal = peserta.jadwal
    except:
        pass

    buffer = buat_kartu_peserta(pendaftaran, jadwal)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="kartu_{pendaftaran.no_pendaftaran}.pdf"'
    )
    return response


@login_required
def cetak_kartu_massal(request):
    """Cetak kartu peserta massal berdasarkan filter"""
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from django.http import HttpResponse
    from seleksi.kartu_pdf import buat_kartu_peserta
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from io import BytesIO

    gelombang_id = request.GET.get('gelombang', '')
    jalur_id     = request.GET.get('jalur', '')
    status       = request.GET.get('status', 'LULUS_ADM')

    qs = Pendaftaran.objects.select_related(
        'user', 'jalur', 'gelombang', 'prodi_pilihan_1'
    )
    if gelombang_id: qs = qs.filter(gelombang_id=gelombang_id)
    if jalur_id:     qs = qs.filter(jalur_id=jalur_id)
    if status:       qs = qs.filter(status=status)

    if not qs.exists():
        messages.warning(request, 'Tidak ada pendaftar yang sesuai filter.')
        return redirect('admin_pmb:pendaftar')

    # Gabungkan semua PDF
    from PyPDF2 import PdfMerger
    merger = PdfMerger()

    for p in qs:
        try:
            buf = buat_kartu_peserta(p)
            merger.append(buf)
            from seleksi.models import KartuPeserta
            KartuPeserta.objects.get_or_create(
                pendaftaran=p,
                defaults={'no_kartu': f'PMB-{p.no_pendaftaran}'}
            )
        except Exception as e:
            logger.error(f'Error cetak kartu {p.no_pendaftaran}: {e}')

    output = BytesIO()
    merger.write(output)
    merger.close()
    output.seek(0)

    response = HttpResponse(output, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="kartu_peserta_massal.pdf"'
    return response

@login_required
def cetak_formulir_admin(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    from django.http import HttpResponse
    from seleksi.kartu_pdf import buat_formulir_pendaftaran

    pendaftaran = get_object_or_404(
        Pendaftaran.objects.select_related(
            'user', 'jalur', 'gelombang',
            'prodi_pilihan_1', 'prodi_pilihan_2'
        ), pk=pk
    )

    buffer = buat_formulir_pendaftaran(pendaftaran)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="formulir_{pendaftaran.no_pendaftaran}.pdf"'
    )
    return response

@login_required
def pembayaran_kwitansi(request, pk):
    if not cek_admin(request.user):
        return redirect('dashboard:index')
    from django.http import HttpResponse, Http404
    from django.shortcuts import get_object_or_404
    from pembayaran.models import KonfirmasiPembayaran
    from pembayaran.pdf import generate_kwitansi_pdf

    konfirmasi = get_object_or_404(
        KonfirmasiPembayaran.objects.filter(status='dikonfirmasi'),
        pk=pk,
    )
    buffer = generate_kwitansi_pdf(konfirmasi)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Kwitansi-{konfirmasi.tagihan.kode_bayar}.pdf"'
    return response

# ========================== LAPORAN PEMBAYARAN ==========================

def _filter_laporan(request):
    """Ambil filter dari GET params & return queryset KonfirmasiPembayaran + metadata."""
    today = date.today()
    
    # Default: 30 hari terakhir
    dari_str = request.GET.get('dari', '')
    sampai_str = request.GET.get('sampai', '')
    jalur_id = request.GET.get('jalur', '')
    gelombang_id = request.GET.get('gelombang', '')
    metode = request.GET.get('metode', '')

    try:
        dari = date.fromisoformat(dari_str) if dari_str else today - timedelta(days=30)
    except ValueError:
        dari = today - timedelta(days=30)
    try:
        sampai = date.fromisoformat(sampai_str) if sampai_str else today
    except ValueError:
        sampai = today

    # Query dasar: hanya konfirmasi yang sudah dikonfirmasi (lunas)
    qs = KonfirmasiPembayaran.objects.filter(
        status='dikonfirmasi',
        tgl_bayar__gte=dari,
        tgl_bayar__lte=sampai,
    ).select_related(
        'tagihan',
        'tagihan__pendaftaran',
        'tagihan__pendaftaran__user',
        'tagihan__pendaftaran__jalur',
        'tagihan__pendaftaran__gelombang',
        'tagihan__pendaftaran__prodi_pilihan_1',
    ).order_by('-tgl_bayar', '-tgl_konfirmasi')

    # Filter opsional
    if jalur_id:
        qs = qs.filter(tagihan__pendaftaran__jalur_id=jalur_id)
    if gelombang_id:
        qs = qs.filter(tagihan__pendaftaran__gelombang_id=gelombang_id)
    if metode:
        qs = qs.filter(metode_bayar=metode)

    return {
        'qs': qs,
        'dari': dari,
        'sampai': sampai,
        'jalur_id': jalur_id,
        'gelombang_id': gelombang_id,
        'metode': metode,
    }


@login_required
def laporan_pembayaran(request):
    """Halaman laporan penerimaan pembayaran dengan filter."""
    if not request.user.is_staff:
        return redirect('dashboard:calon_maba')

    filters = _filter_laporan(request)
    qs = filters['qs']

    # Ringkasan
    total_count = qs.count()
    total_amount = qs.aggregate(t=Sum('jumlah_bayar'))['t'] or 0

    # Breakdown per metode
    breakdown = qs.values('metode_bayar').annotate(
        count=Count('id'),
        total=Sum('jumlah_bayar'),
    ).order_by('-total')

    # Data tambahan untuk filter dropdown
    from master.models import JalurPenerimaan, GelombangPenerimaan
    jalur_list = JalurPenerimaan.objects.filter(status='aktif').order_by('urutan', 'nama_jalur')
    gelombang_list = GelombangPenerimaan.objects.select_related('jalur').order_by('-tahun_akademik', 'tgl_buka')

    context = {
        'konfirmasi_list': qs,
        'total_count': total_count,
        'total_amount': total_amount,
        'breakdown': breakdown,
        'dari': filters['dari'],
        'sampai': filters['sampai'],
        'jalur_id': filters['jalur_id'],
        'gelombang_id': filters['gelombang_id'],
        'metode': filters['metode'],
        'jalur_list': jalur_list,
        'gelombang_list': gelombang_list,
        'METODE_CHOICES': KonfirmasiPembayaran.METODE_CHOICES,
    }
    return render(request, 'admin_pmb/laporan_pembayaran.html', context)


@login_required
def laporan_pembayaran_excel(request):
    """Export Excel: Ringkasan + Detail."""
    if not request.user.is_staff:
        return redirect('dashboard:calon_maba')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    filters = _filter_laporan(request)
    qs = filters['qs']
    dari = filters['dari']
    sampai = filters['sampai']

    wb = openpyxl.Workbook()

    # ===== Sheet 1: Ringkasan =====
    ws1 = wb.active
    ws1.title = "Ringkasan"

    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2563EB')
    total_font = Font(bold=True, size=11)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1['A1'] = 'LAPORAN PENERIMAAN PEMBAYARAN PMB UNISAN'
    ws1['A1'].font = header_font
    ws1['A1'].fill = header_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.merge_cells('A1:D1')
    ws1.row_dimensions[1].height = 28

    ws1['A2'] = f'Periode: {dari.strftime("%d %b %Y")}  s/d  {sampai.strftime("%d %b %Y")}'
    ws1.merge_cells('A2:D2')
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1['A2'].font = Font(italic=True, color='6B7280')

    # Total
    ws1['A4'] = 'Total Transaksi:'
    ws1['B4'] = qs.count()
    ws1['A5'] = 'Total Nominal:'
    ws1['B5'] = float(qs.aggregate(t=Sum('jumlah_bayar'))['t'] or 0)
    ws1['B5'].number_format = '"Rp "#,##0'
    ws1['A4'].font = total_font
    ws1['A5'].font = total_font

    # Breakdown per metode
    ws1['A7'] = 'BREAKDOWN PER METODE'
    ws1['A7'].font = Font(bold=True, size=12)
    ws1.merge_cells('A7:D7')

    row = 8
    headers = ['Metode', 'Jumlah Transaksi', 'Total Nominal', '%']
    for col, h in enumerate(headers, start=1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='E5E7EB')
        cell.border = border

    total_for_pct = float(qs.aggregate(t=Sum('jumlah_bayar'))['t'] or 1)
    METODE_DISPLAY = dict(KonfirmasiPembayaran.METODE_CHOICES)
    for b in qs.values('metode_bayar').annotate(c=Count('id'), t=Sum('jumlah_bayar')).order_by('-t'):
        row += 1
        total = float(b['t'] or 0)
        ws1.cell(row=row, column=1, value=METODE_DISPLAY.get(b['metode_bayar'], b['metode_bayar'])).border = border
        ws1.cell(row=row, column=2, value=b['c']).border = border
        c = ws1.cell(row=row, column=3, value=total)
        c.number_format = '"Rp "#,##0'
        c.border = border
        pct = (total / total_for_pct * 100) if total_for_pct else 0
        c = ws1.cell(row=row, column=4, value=f"{pct:.1f}%")
        c.alignment = Alignment(horizontal='right')
        c.border = border

    # Width
    for col, w in zip('ABCD', [22, 18, 22, 10]):
        ws1.column_dimensions[col].width = w

    # ===== Sheet 2: Detail =====
    ws2 = wb.create_sheet('Detail Transaksi')
    cols = ['No', 'Tgl Bayar', 'Kode Tagihan', 'Nama Pendaftar', 'No Pendaftaran',
            'Jalur', 'Gelombang', 'Prodi', 'Metode', 'Bank Asal', 'Atas Nama',
            'No Transaksi', 'Jumlah', 'Dikonfirmasi']

    for col, h in enumerate(cols, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2563EB')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for i, k in enumerate(qs, start=1):
        row = i + 1
        p = k.tagihan.pendaftaran
        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=k.tgl_bayar)
        ws2.cell(row=row, column=2).number_format = 'DD-MM-YYYY'
        ws2.cell(row=row, column=3, value=k.tagihan.kode_bayar)
        ws2.cell(row=row, column=4, value=p.user.get_full_name() or p.user.username)
        ws2.cell(row=row, column=5, value=p.no_pendaftaran)
        ws2.cell(row=row, column=6, value=p.jalur.nama if p.jalur else '-')
        ws2.cell(row=row, column=7, value=p.gelombang.nama if p.gelombang else '-')
        ws2.cell(row=row, column=8, value=p.prodi_pilihan_1.nama_prodi if p.prodi_pilihan_1 else '-')
        ws2.cell(row=row, column=9, value=METODE_DISPLAY.get(k.metode_bayar, k.metode_bayar))
        ws2.cell(row=row, column=10, value=k.bank_asal or '-')
        ws2.cell(row=row, column=11, value=k.atas_nama_pengirim or '-')
        ws2.cell(row=row, column=12, value=k.no_transaksi or '-')
        c = ws2.cell(row=row, column=13, value=float(k.jumlah_bayar))
        c.number_format = '"Rp "#,##0'
        ws2.cell(row=row, column=14, value=k.tgl_konfirmasi.strftime('%d-%m-%Y %H:%M') if k.tgl_konfirmasi else '-')

        for col in range(1, 15):
            ws2.cell(row=row, column=col).border = border

    # Total row
    last_row = qs.count() + 2
    ws2.cell(row=last_row, column=12, value='TOTAL').font = Font(bold=True)
    ws2.cell(row=last_row, column=12).alignment = Alignment(horizontal='right')
    c = ws2.cell(row=last_row, column=13, value=float(qs.aggregate(t=Sum('jumlah_bayar'))['t'] or 0))
    c.number_format = '"Rp "#,##0'
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='FEF3C7')

    # Auto-width
    widths = [5, 12, 18, 28, 18, 14, 14, 25, 14, 12, 22, 20, 15, 18]
    for col, w in zip('ABCDEFGHIJKLMN', widths):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = 'A2'

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Laporan_Pembayaran_{dari.strftime("%Y%m%d")}_{sampai.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def laporan_pembayaran_pdf(request):
    """Export PDF laporan — A4 landscape dengan TTD bendahara."""
    if not request.user.is_staff:
        return redirect('dashboard:calon_maba')

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
    from reportlab.lib.utils import ImageReader
    from django.http import HttpResponse
    from django.utils import timezone
    import os
    from django.conf import settings

    filters = _filter_laporan(request)
    qs = filters['qs']
    dari = filters['dari']
    sampai = filters['sampai']

    response = HttpResponse(content_type='application/pdf')
    filename = f'Laporan_Pembayaran_{dari.strftime("%Y%m%d")}_{sampai.strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.2*cm, bottomMargin=1.2*cm,
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('title', parent=styles['Title'],
                                  fontName='Helvetica-Bold', fontSize=14,
                                  alignment=TA_CENTER, spaceAfter=4)
    subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'],
                                     fontName='Helvetica', fontSize=10,
                                     alignment=TA_CENTER, textColor=colors.grey)

    story.append(Paragraph("LAPORAN PENERIMAAN PEMBAYARAN PMB UNISAN", title_style))
    story.append(Paragraph(
        f"Periode: {dari.strftime('%d %B %Y')} s/d {sampai.strftime('%d %B %Y')}",
        subtitle_style
    ))
    story.append(Spacer(1, 0.4*cm))

    # Ringkasan box
    total_count = qs.count()
    total_amount = float(qs.aggregate(t=Sum('jumlah_bayar'))['t'] or 0)

    ringkasan_data = [
        ['Total Transaksi', f'{total_count:,}', 'Total Nominal', f'Rp {total_amount:,.0f}'.replace(',', '.')],
    ]
    t = Table(ringkasan_data, colWidths=[4*cm, 4*cm, 4*cm, 6*cm])
    t.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F3F4F6')),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (2, 0), (2, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('FONT', (3, 0), (3, -1), 'Helvetica-Bold', 11),
        ('TEXTCOLOR', (3, 0), (3, -1), colors.HexColor('#059669')),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Tabel detail
    header = ['No', 'Tgl Bayar', 'Kode', 'Pendaftar', 'Jalur / Gel.', 'Metode', 'Jumlah']
    data = [header]

    METODE_DISPLAY = dict(KonfirmasiPembayaran.METODE_CHOICES)
    for i, k in enumerate(qs, start=1):
        p = k.tagihan.pendaftaran
        nama = (p.user.get_full_name() or p.user.username)[:28]
        jalur_gel = f"{p.jalur.nama[:8] if p.jalur else '-'} / {p.gelombang.nama[:14] if p.gelombang else '-'}"
        data.append([
            str(i),
            k.tgl_bayar.strftime('%d-%m-%Y') if k.tgl_bayar else '-',
            k.tagihan.kode_bayar,
            nama,
            jalur_gel,
            METODE_DISPLAY.get(k.metode_bayar, k.metode_bayar),
            f'Rp {float(k.jumlah_bayar):,.0f}'.replace(',', '.'),
        ])

    # Total row
    data.append(['', '', '', '', '', 'TOTAL', f'Rp {total_amount:,.0f}'.replace(',', '.')])

    table = Table(data, colWidths=[1*cm, 2.5*cm, 3.5*cm, 6*cm, 5*cm, 3*cm, 3.5*cm], repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Body
        ('FONT', (0, 1), (-1, -2), 'Helvetica', 8.5),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F9FAFB')]),
        # Total
        ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 9.5),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.grey),
        # Grid
        ('GRID', (0, 0), (-1, -2), 0.25, colors.HexColor('#E5E7EB')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    # TTD Bendahara
    story.append(Spacer(1, 0.8*cm))

    try:
        from master.models import PengaturanSistem
        pengaturan = PengaturanSistem.get()
    except Exception:
        pengaturan = None

    nama_bendahara = (pengaturan.nama_bendahara_pmb if pengaturan else '') or 'Bendahara Panitia PMB'
    nip_bendahara = (pengaturan.nip_bendahara_pmb if pengaturan else '') or ''
    ttd_img_path = None
    if pengaturan and pengaturan.ttd_bendahara_pmb:
        try:
            if os.path.exists(pengaturan.ttd_bendahara_pmb.path):
                ttd_img_path = pengaturan.ttd_bendahara_pmb.path
        except Exception:
            pass

    today_str = timezone.localdate().strftime('%d %B %Y')
    ttd_content = [
        Paragraph(f'Gorontalo, {today_str}', ParagraphStyle(
            'r', parent=styles['Normal'], fontSize=10, alignment=TA_RIGHT
        )),
        Paragraph('Bendahara Panitia PMB UNISAN', ParagraphStyle(
            'r2', parent=styles['Normal'], fontSize=10, alignment=TA_RIGHT
        )),
    ]
    if ttd_img_path:
        ttd_content.append(Spacer(1, 0.2*cm))
        img = Image(ttd_img_path, width=4*cm, height=1.8*cm)
        img.hAlign = 'RIGHT'
        ttd_content.append(img)
    else:
        ttd_content.append(Spacer(1, 2*cm))

    ttd_content.append(Paragraph(
        f'<b>{nama_bendahara}</b>',
        ParagraphStyle('r3', parent=styles['Normal'], fontSize=10, alignment=TA_RIGHT)
    ))
    if nip_bendahara:
        ttd_content.append(Paragraph(
            f'NIP. {nip_bendahara}',
            ParagraphStyle('r4', parent=styles['Normal'], fontSize=9, alignment=TA_RIGHT, textColor=colors.grey)
        ))

    ttd_table = Table(
        [['', ttd_content]],
        colWidths=[14*cm, 10*cm],
    )
    ttd_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(ttd_table)

    doc.build(story)
    return response

# ============================================================
# SETUP PRODI per GELOMBANG (Bulk Matrix Editor)
# ============================================================

from master.models import ProdiPMB
from master.services.setup_prodi import (
    get_matrix_prodi,
    save_matrix_prodi,
    clone_prodi_gelombang,
    get_gelombang_with_count,
)


@login_required
def setup_prodi_list(request):
    """List semua gelombang dengan jumlah prodi yang sudah disetup."""
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    gelombang_list = get_gelombang_with_count()

    # Hitung total prodi master untuk reference progress
    from utils.simda_reader import get_program_studi
    total_prodi_master = len(get_program_studi())

    return render(request, 'admin_pmb/setup_prodi_list.html', {
        **get_sidebar_counts(),
        'gelombang_list':     gelombang_list,
        'total_prodi_master': total_prodi_master,
    })


@login_required
def setup_prodi_matrix(request, gelombang_id):
    """Matrix bulk edit prodi untuk satu gelombang."""
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    gelombang = get_object_or_404(GelombangPenerimaan, pk=gelombang_id)

    if request.method == 'POST':
        # Parse data form
        kode_list = request.POST.getlist('kode_prodi[]')
        rows_data = []

        for kode in kode_list:
            is_checked = request.POST.get(f'check_{kode}') == 'on'
            rows_data.append({
                'kode_prodi':    kode,
                'nama_prodi':    request.POST.get(f'nama_prodi_{kode}', ''),
                'kode_fakultas': request.POST.get(f'kode_fakultas_{kode}', ''),
                'nama_fakultas': request.POST.get(f'nama_fakultas_{kode}', ''),
                'is_checked':    is_checked,
                'kuota':         request.POST.get(f'kuota_{kode}', 0),
                'daya_tampung':  request.POST.get(f'daya_tampung_{kode}', 0),
                'biaya_kuliah':  request.POST.get(f'biaya_kuliah_{kode}', 0),
                'biaya_spp':     request.POST.get(f'biaya_spp_{kode}', 0),
            })

        try:
            counts = save_matrix_prodi(gelombang, rows_data)
            messages.success(
                request,
                f"Berhasil disimpan: {counts['created']} ditambah, "
                f"{counts['updated']} diupdate, {counts['deleted']} dihapus."
            )
        except Exception as e:
            logger.exception("Gagal save matrix prodi")
            messages.error(request, f"Gagal menyimpan: {e}")

        return redirect('admin_pmb:setup_prodi_matrix', gelombang_id=gelombang.id)

    matrix = get_matrix_prodi(gelombang)

    # Untuk Clone dropdown — gelombang lain yang punya minimal 1 prodi
    gelombang_sumber = GelombangPenerimaan.objects.exclude(pk=gelombang.pk)\
        .filter(prodi_pmb__isnull=False).distinct()\
        .select_related('jalur').order_by('-tahun_akademik', '-tgl_buka')

    # Stats
    total_checked = sum(1 for r in matrix if r['is_checked'])
    total_master = len(matrix)

    return render(request, 'admin_pmb/setup_prodi_matrix.html', {
        **get_sidebar_counts(),
        'gelombang':         gelombang,
        'matrix':            matrix,
        'gelombang_sumber':  gelombang_sumber,
        'total_checked':     total_checked,
        'total_master':      total_master,
    })


@login_required
def setup_prodi_clone(request, gelombang_id):
    """Action clone prodi dari gelombang sumber ke target."""
    if not cek_admin(request.user):
        return redirect('dashboard:index')

    if request.method != 'POST':
        return redirect('admin_pmb:setup_prodi_matrix', gelombang_id=gelombang_id)

    target = get_object_or_404(GelombangPenerimaan, pk=gelombang_id)
    source_id = request.POST.get('source_gelombang_id')

    if not source_id:
        messages.error(request, "Pilih gelombang sumber terlebih dahulu.")
        return redirect('admin_pmb:setup_prodi_matrix', gelombang_id=gelombang_id)

    source = get_object_or_404(GelombangPenerimaan, pk=source_id)

    if source.pk == target.pk:
        messages.error(request, "Gelombang sumber dan target tidak boleh sama.")
        return redirect('admin_pmb:setup_prodi_matrix', gelombang_id=gelombang_id)

    try:
        counts = clone_prodi_gelombang(source, target)
        messages.success(
            request,
            f"Berhasil clone dari {source.nama_gelombang}: "
            f"{counts['created']} ditambah, {counts['updated']} diupdate."
        )
    except Exception as e:
        logger.exception("Gagal clone prodi")
        messages.error(request, f"Gagal clone: {e}")

    return redirect('admin_pmb:setup_prodi_matrix', gelombang_id=gelombang_id)